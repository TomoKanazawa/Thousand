from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Outreach"

headers = [
    "Contact", "Institution", "Email", "Paper / Study",
    "Their fit", "Date sent", "Replied?", "Reply date",
    "Status / Next step", "Notes",
]
ws.append(headers)

rows = [
    [
        "Adrian D. Haimovich, MD, PhD",
        "Beth Israel Deaconess / Harvard",
        "ahaimovi@bidmc.harvard.edu",
        "Haimovich 2025 — eTriggers + Claude (medRxiv)",
        "LLM + Safer Dx, ED",
        "",
        "Yes",
        "",
        "In conversation — keep warm",
        "Same Boston cluster as Dalal; potential warm-intro path",
    ],
    [
        "Katie Raffel, MD",
        "University of Colorado Anschutz",
        "katie.raffel@cuanschutz.edu",
        "Raffel 2020 — 7-day readmission dx-errors (BMJ Qual Saf)",
        "Inpatient + named final dx + Safer Dx",
        "2026-06-06",
        "Pending",
        "",
        "Awaiting reply",
        "Best dataset fit; warm-intro path to Auerbach (UCSF/UPSIDE colleague)",
    ],
]
for r in rows:
    ws.append(r)

# Header style
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="305496")
header_align = Alignment(horizontal="left", vertical="center")
for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align

# Body font + wrap for long cells
body_font = Font(name="Arial")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for c in row:
        c.font = body_font
        c.alignment = Alignment(vertical="top", wrap_text=True)

# Column widths
widths = {
    "A": 30,  # Contact
    "B": 32,  # Institution
    "C": 32,  # Email
    "D": 42,  # Paper
    "E": 30,  # Fit
    "F": 13,  # Date sent
    "G": 12,  # Replied?
    "H": 12,  # Reply date
    "I": 30,  # Status
    "J": 50,  # Notes
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze header
ws.freeze_panes = "A2"

# Conditional formatting on Replied? column (G)
green = PatternFill("solid", start_color="C6EFCE")
yellow = PatternFill("solid", start_color="FFEB9C")
red = PatternFill("solid", start_color="FFC7CE")
rng = f"G2:G1000"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Yes"'], fill=green))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pending"'], fill=yellow))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"No"'], fill=red))

# Row height bump
ws.row_dimensions[1].height = 22
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 42

wb.save("/Users/tomohiro/Projects/Thousand/outreach/outreach_tracker.xlsx")
print("OK")
